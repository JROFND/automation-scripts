import shutil
import os
import re
import pandas as pd
import os
import glob
import time
import openpyxl
from pathlib import Path
from openpyxl import load_workbook
import numpy as np
from datetime import datetime, date

# move files from downloads to walgreens input folder that were
# modified today and contain the keyword in the filename
source_directory = 'C:/Users/703341565/Downloads' # Replace with the actual source directory
downloads_directory = "W:/SHARE2 Pharmacy Department Share/340B/Jeff R/Walgreens/Input"  # Replace with the actual downloads directory
keyword = "340BComplete_RxDetails"  # Replace with the keyword you're looking for
today = date.today()
for filename in os.listdir(source_directory):
    filepath = os.path.join(source_directory, filename)
    if os.path.isfile(filepath):
                mtime = datetime.fromtimestamp(os.path.getmtime(filepath)).date()
                destination_filepath = os.path.join(downloads_directory, filename)
                if mtime == today and keyword in filename and filename.endswith('.xlsx'):
                    files = shutil.copy2(filepath, destination_filepath)  
                    # excel_file = pd.ExcelFile(excel_files)
                    # ret_df=[]
                    # spec_df = []
                    # for e in files:                        
                    workbook = openpyxl.load_workbook(files)
                    sheet = workbook.active                    
                    columns_to_numeric = ['Plan AR Amt','Copay Amt','Sales Tax','Admin Fee','Dispense Fee','Insured Admin Fee','Qty Disp.','Estimated 340B Unit Price','Estimated 340B Cost','Est. AWP Unit Price','Est. AWP Cost']
                    columns_to_date = ['Rx Written Date','Adjud. Date','Fill Sold Date','Reversal Date','Date sent to ESP']
                    rows_to_remove = [0,1,2,3,4,5,6]
                    ret_df=[]
                    # spec_df = []
                    keyword = "Retail"   
                    matching_sheets = [sheet for sheet in workbook.sheetnames if keyword.lower() in sheet.lower()]   
                    for m in matching_sheets:           
                        new_filename_base =sheet["E4"].value.replace("Entity Name","Walgreens")+sheet["E3"].value.replace("Time Period","_Time_Period")
                        # # Check if the cell has a value
                        if new_filename_base:
                            # Sanitize the filename
                            valid_filename = "".join(c if c.isalnum() or c in (' ', '.', '_') else '_' for c in str(new_filename_base)).strip()
                            new_filename = f"{valid_filename}.xlsx"
                            retdf = pd.read_excel(new_filename,sheet_name=sheet, header=None, 
                            names =('Patient Name', 'Patient Birthdate', 'Prescriber Name','Prescriber NPI', 'Pharmacy Intersection',
                            'Pharmacy #', 'Pharmacy NPI', 'Rx Nbr', 'Fill Nbr', 
                            'Rx Source', 'Rx Written Date', 'Adjud. Date', 
                            'Fill Sold Date', 'Original Billing Month/Year', 
                            'Claim Type', 'Payer Type', 'Group Id', 
                            'Group Name', 'Uninsured Cardholder ID', 
                            'EMR Member ID', 'NDC 11', 'Drug Name',
                            'Manufacturer Name', 'Therapeutic Class', 'Drug Class', 
                            'Sold/Reversals', 'Reversal Date', 'Qty Disp.', 
                            'Days Supply', 'Estimated 340B Unit Price', 'Estimated 340B Cost', 
                            'Plan AR Amt','Copay Amt', 'Sales Tax', 'Admin Fee',
                            'Dispense Fee', 'Insured Admin Fee', 'Increased Access Dollars ',
                            'Est. AWP Unit Price', 'Est. AWP Cost',  'Location ID', 
                            'HCS ID','Client Name', 'Clinic Name', 
                            'Department Name', 'Order ID', 'Referral Case Number', 
                            'Date sent to ESP','Time_Period','Entity'))
                            ret_df.append(retdf) 
                            
                            
                            
                            
# from sqlalchemy import Table, Column
# from sqlalchemy import create_engine
# from sqlalchemy import types
# from sqlalchemy import MetaData
# from sqlalchemy import text

# pharmserver = 'SQLSERVER6'
# pharmdb = 'PharmMart'
# pharmconn = create_engine('mssql+pyodbc://@' + pharmserver + '/' +
#                           pharmdb + '?trusted_connection=yes&driver=ODBC+Driver+17+for+SQL+Server')
                 
        # Get the modification time and convert it to a date

        
        

#                 # Check if the file was modified today and contains the keyword
#         

#             shutil.copy2(filepath, destination_filepath)  # Use copy2 to preserve metadata
#             print(f"Copied '{filename}' to '{downloads_directory}'")


# folder_path = r'W://SHARE2 Pharmacy Department Share//340B//Jeff R//Walgreens//Input//'
# os.chdir(folder_path)
# all_files = os.listdir(folder_path)

# excel_files = glob.glob(folder_path+'Walgreens*.xlsx')
# columns_to_numeric = ['Plan AR Amt','Copay Amt','Sales Tax','Admin Fee','Dispense Fee','Insured Admin Fee','Qty Disp.','Estimated 340B Unit Price','Estimated 340B Cost','Est. AWP Unit Price','Est. AWP Cost']
# columns_to_date = ['Rx Written Date','Adjud. Date','Fill Sold Date','Reversal Date','Date sent to ESP']

# files = [f for f in all_files if f.startswith('Sanford')]
# excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx') or f.endswith('.xls')]
# rows_to_remove = [0,1,2,3,4,5,6]


# # # # for file in excel_files:
# for e in excel_files:
#     excel_file = pd.ExcelFile(e)
#     # empty_df = pd.DataFrame()
#     ret_df=[]
#     spec_df = []
#     for sheet in excel_file.sheet_names:
#         if "Retail" in sheet:
#             retdf = pd.read_excel(e,sheet_name=sheet, header=None, names =('Patient Name', 'Patient Birthdate', 'Prescriber Name','Prescriber NPI', 'Pharmacy Intersection', 'Pharmacy #', 'Pharmacy NPI', 'Rx Nbr', 'Fill Nbr', 
#             'Rx Source', 'Rx Written Date', 'Adjud. Date', 'Fill Sold Date', 'Original Billing Month/Year', 'Claim Type', 'Payer Type', 'Group Id', 'Group Name', 'Uninsured Cardholder ID', 'EMR Member ID', 
#             'NDC 11', 'Drug Name', 'Manufacturer Name', 'Therapeutic Class', 'Drug Class', 'Sold/Reversals', 'Reversal Date', 'Qty Disp.', 'Days Supply', 'Estimated 340B Unit Price', 'Estimated 340B Cost', 
#             'Plan AR Amt','Copay Amt', 'Sales Tax', 'Admin Fee', 'Dispense Fee', 'Insured Admin Fee', 'Increased Access Dollars ', 'Est. AWP Unit Price', 'Est. AWP Cost', 
#             'Location ID', 'HCS ID','Client Name', 'Clinic Name', 'Department Name', 'Order ID', 'Referral Case Number', 
#             'Date sent to ESP','Time_Period','Entity'))
#             retdf['Time_Period'] =  retdf.iloc[2, 4].replace("Time Period:","")
#             retdf['Entity'] =  retdf.iloc[3, 4].replace("Entity Name:","")     
#             retdf = retdf.drop(rows_to_remove) 
#             ret_df.append(retdf)  
            
#             retail = pd.concat(ret_df)       
#             retail[columns_to_numeric] = retail[columns_to_numeric].replace('', np.nan).fillna(0).apply(pd.to_numeric, errors='coerce').round(2)  
#             retail['Estimated 340B Cost'] = retail['Estimated 340B Cost']* -1           
#             retail[columns_to_date] = retail[columns_to_date].apply(pd.to_datetime, errors='coerce') 
#             retail['Total Revenue'] = ((retail['Plan AR Amt'] + retail['Copay Amt'])-retail['Sales Tax'])
#             retail['Total Dispense Fee'] = retail['Admin Fee'] + retail['Dispense Fee'] + retail['Insured Admin Fee']
#             retail['Profit'] =  retail['Total Revenue'] - retail['Total Dispense Fee'] - retail['Estimated 340B Cost']
#             # retail.to_sql('340B_Enterprise_Walgreens_RX_Retail',con=pharmconn, if_exists='replace',chunksize=1000, 
#             #                                                   method=None,index=False)
        
        
#         elif "Specialty" in sheet:
#             specdf = pd.read_excel(e,sheet_name=sheet, header=None, 
#              names =('Patient Name', 'Patient Birthdate', 'Prescriber Name','Prescriber NPI', 'Pharmacy Intersection', 'Pharmacy #', 'Pharmacy NPI', 'Rx Nbr', 'Fill Nbr', 
#             'Rx Source', 'Rx Written Date', 'Adjud. Date', 'Fill Sold Date', 'Original Billing Month/Year', 'Claim Type', 'Payer Type', 'Group Id', 'Group Name', 'Uninsured Cardholder ID', 'EMR Member ID', 
#             'NDC 11', 'Drug Name', 'Manufacturer Name', 'Therapeutic Class', 'Drug Class', 'Sold/Reversals', 'Reversal Date', 'Qty Disp.', 'Days Supply', 'Estimated 340B Unit Price', 'Estimated 340B Cost', 
#             'Plan AR Amt','Copay Amt', 'Sales Tax', 'Admin Fee', 'Dispense Fee', 'Insured Admin Fee', 'Increased Access Dollars ', 'Est. AWP Unit Price', 'Est. AWP Cost', 
#             'Location ID', 'HCS ID','Client Name', 'Clinic Name', 'Department Name', 'Order ID', 'Referral Case Number', 
#             'Date sent to ESP','Time_Period','Entity'))
#             specdf['Time_Period'] =  specdf.iloc[2, 4].replace("Time Period:","")
#             specdf['Entity'] =  specdf.iloc[3, 4].replace("Entity Name:","")     
#             specdf = specdf.drop(rows_to_remove) 
#             spec_df.append(specdf)  
            
        
#             specialty = pd.concat(spec_df)       
#             specialty[columns_to_numeric] = retail[columns_to_numeric].replace('', np.nan).fillna(0).apply(pd.to_numeric, errors='coerce').round(2)  
#             specialty['Estimated 340B Cost'] = specialty['Estimated 340B Cost']* -1           
#             specialty[columns_to_date] = specialty[columns_to_date].apply(pd.to_datetime, errors='coerce') 
#             specialty['Total Revenue'] = ((specialty['Plan AR Amt'] + specialty['Copay Amt'])-specialty['Sales Tax'])
#             specialty['Total Dispense Fee'] = specialty['Admin Fee'] + retail['Dispense Fee'] + specialty['Insured Admin Fee']
#             specialty['Profit'] =  specialty['Total Revenue'] - specialty['Total Dispense Fee'] - specialty['Estimated 340B Cost']
#             # specialty.to_sql('340B_Enterprise_Walgreens_RX_Specialty',con=pharmconn, if_exists='replace',chunksize=1000, 
            #                                                  method=None,index=False)
